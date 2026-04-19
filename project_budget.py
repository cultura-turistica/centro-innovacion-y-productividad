import os
import openpyxl

def main():
    file_path = os.path.expanduser("~/Downloads/Matriz PDT Chairá VF 1.xlsx")
    out_path = os.path.expanduser("~/Downloads/Matriz PDT Chairá VF 1_Proyectada.xlsx")

    print(f"Loading '{file_path}'...")
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    wb_write = openpyxl.load_workbook(file_path)

    # Find sheets
    a1_sheet_name = [s for s in wb_data.sheetnames if "ANEXO 1" in s][0]
    a2_sheet_name = [s for s in wb_data.sheetnames if "ANEXO 2" in s][0]

    ws_a1_data = wb_data[a1_sheet_name]
    ws_a2_write = wb_write[a2_sheet_name]

    print("Extracting budgets from ANEXO 1...")
    budgets = {}
    for row in range(1, ws_a1_data.max_row + 1):
        act_val = ws_a1_data.cell(row=row, column=9).value
        cost_val = ws_a1_data.cell(row=row, column=10).value
        if isinstance(act_val, str) and act_val.strip() != "":
            if not isinstance(cost_val, (int, float)) and isinstance(cost_val, str) and cost_val.replace('.', '', 1).isdigit():
                try: cost_val = float(cost_val)
                except: pass
            
            if isinstance(cost_val, (int, float)):
                act = act_val.strip().replace('\n', '')
                # Anexo 1 está en PESOS ABSOLUTOS, Anexo 2 requiere cifra en MILLONES.
                budgets[act] = cost_val / 1_000_000.0

    print(f"Found {len(budgets)} activities with budgets.")

    # Base curve peaking at year 7 and 8 (used for continuous operations like marketing, admin, etc.)
    # 10 years curve percentages: Low start, high middle-late, low end
    curve_continuous = [0.03, 0.04, 0.06, 0.09, 0.12, 0.15, 0.20, 0.17, 0.09, 0.05]

    def get_90_split(act_text, val_90):
        text = act_text.lower()
        split = [None]*6 # SGR, FONTUR, DES, INT, MIX, CEN
        if any(w in text for w in ["construcción", "infraestructura", "obra", "vías", "saneamiento", "muelle", "dotación"]):
            split[0] = val_90 # SGR
        elif any(w in text for w in ["promoción", "marca", "marketing", "mercadeo", "competitividad", "diseño", "promocional"]):
            split[1] = val_90 # FONTUR
        elif any(w in text for w in ["comunidad", "paz", "social", "capacitación", "sensibilización", "taller", "cultural"]):
            split[2] = val_90 # DES
        elif any(w in text for w in ["ciencia", "estudio", "tecnología", "información", "monitoreo", "software", "estadístic", "observatorio"]):
            split[5] = val_90 # CEN
        else:
            split[4] = val_90 # MIX
        return split

    print("Processing ANEXO 2...")
    total_budget_anexo2 = 0
    total_per_year = [0]*10

    for row in range(5, ws_a2_write.max_row + 1):
        act_cell = ws_a2_write.cell(row=row, column=3).value
        
        if isinstance(act_cell, str) and act_cell.strip() != "":
            act_clean = act_cell.strip().replace('\n', '')
            total = budgets.get(act_clean)
            if total is None:
                for k, v in budgets.items():
                    if k.replace(" ", "") == act_clean.replace(" ", ""):
                        total = v
                        break

            if total is not None:
                total_budget_anexo2 += total
                
                # Determine if fixed/infrastructure or continuous
                text_lower = act_clean.lower()
                is_mega_infra = any(w in text_lower for w in ["construcción", "infraestructura", "malecón", "muelle", "pavimentación"])
                is_fixed_soft = any(w in text_lower for w in ["formulación", "diseño", "marca", "política pública", "adopción", "creación de la marca"])
                
                distribution = [0]*10
                
                if is_mega_infra and total > 500: # > 500 Millones
                    # Split into phases: Year 5 (Feasibility: 5%), Year 6 (Permits/Start: 25%), Year 7 (Peak Build: 45%), Year 8 (Finish/Delivery: 25%)
                    distribution[4] = total * 0.05
                    distribution[5] = total * 0.25
                    distribution[6] = total * 0.45
                    distribution[7] = total * 0.25
                elif is_mega_infra and total > 100: # Medium Infra > 100 Millones
                    # Year 6 (20%), Year 7 (50%), Year 8 (30%)
                    distribution[5] = total * 0.20
                    distribution[6] = total * 0.50
                    distribution[7] = total * 0.30
                elif is_fixed_soft:
                    # Early planning stuff like Policies -> phase Year 1-3
                    # Brand creation -> Year 3-4
                    if "marca" in text_lower or "identidad" in text_lower:
                        distribution[2] = total * 0.40 # Year 3
                        distribution[3] = total * 0.60 # Year 4
                    else: # Policy / Formulation
                        distribution[0] = total * 0.30 # Year 1
                        distribution[1] = total * 0.60 # Year 2
                        distribution[2] = total * 0.10 # Year 3
                else:
                    # Continuous programs, capacity building, events, etc.
                    distribution = [total * c for c in curve_continuous]

                for y in range(10):
                    val = distribution[y]
                    if val > 0.001: # Representa al menos 1,000 pesos
                        total_per_year[y] += val
                        mu = val * 0.10
                        resto = val * 0.90
                        splits = get_90_split(act_clean, resto)
                        
                        base_col = 5 + y*7
                        ws_a2_write.cell(row=row, column=base_col).value = mu # MU
                        if splits[0]: ws_a2_write.cell(row=row, column=base_col+1).value = splits[0] # SGR
                        if splits[1]: ws_a2_write.cell(row=row, column=base_col+2).value = splits[1] # FONTUR
                        if splits[2]: ws_a2_write.cell(row=row, column=base_col+3).value = splits[2] # DES
                        if splits[3]: ws_a2_write.cell(row=row, column=base_col+4).value = splits[3] # INT
                        if splits[4]: ws_a2_write.cell(row=row, column=base_col+5).value = splits[4] # MIX
                        if splits[5]: ws_a2_write.cell(row=row, column=base_col+6).value = splits[5] # CEN
            
    print(f"Total budget applied: ${total_budget_anexo2:,.2f}")
    print("Yearly distribution:")
    for y, amount in enumerate(total_per_year):
        print(f"Año {y+1}: ${amount:,.2f} ({(amount/total_budget_anexo2)*100:.1f}%)")
    
    # Save file
    print(f"Saving to '{out_path}'...")
    wb_write.save(out_path)
    print("File saved successfully!")

if __name__ == "__main__":
    main()
